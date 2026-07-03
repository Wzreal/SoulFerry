import smtplib
import ssl
import threading
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.core.enums import ToolStatus
from app.models.entities import AlertRecord, ExcelRecord, PsychologicalReport, UserAccount


EXCEL_WRITE_LOCK = threading.Lock()


class ToolOrchestrationService:
    def __init__(self, db: Session, settings: Settings):
        self.db = db
        self.settings = settings

    def write_excel(self, report: PsychologicalReport) -> ExcelRecord:
        existing = (
            self.db.query(ExcelRecord)
            .filter(ExcelRecord.report_id == report.id, ExcelRecord.status == ToolStatus.SUCCESS.value)
            .first()
        )
        if existing is not None:
            return existing
        path = Path(self.settings.excel_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with EXCEL_WRITE_LOCK:
            if path.exists():
                workbook = load_workbook(path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "SoulFerry Risk Ledger"
                sheet.append(["reportId", "riskLevel", "emotion", "confidence", "summary", "createdAt"])
            sheet.append([report.id, report.risk_level, report.emotion, report.confidence, report.summary, report.created_at.isoformat()])
            workbook.save(path)
        record = ExcelRecord(report_id=report.id, file_path=str(path), status=ToolStatus.SUCCESS.value, message="Excel 台账已写入")
        self.db.add(record)
        self.db.commit()
        return record

    def notify(self, report: PsychologicalReport) -> AlertRecord:
        existing = (
            self.db.query(AlertRecord)
            .filter(AlertRecord.report_id == report.id, AlertRecord.status == ToolStatus.SUCCESS.value)
            .first()
        )
        if existing is not None:
            return existing
        recipient = self.settings.alert_email_to.strip() or "unconfigured"
        mode = self.settings.alert_email_delivery_mode.strip().lower()
        if mode == "log":
            return self._save_alert(
                report,
                recipient if recipient != "unconfigured" else "log",
                ToolStatus.SUCCESS.value,
                f"高风险预警已记录：reportId={report.id}，deliveryMode=log",
            )
        if mode != "smtp":
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件未发送：未知投递模式 {self.settings.alert_email_delivery_mode}",
            )
        missing = self._missing_email_config()
        if missing:
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件未发送：缺少配置 {', '.join(missing)}",
            )
        try:
            self._send_alert_email(report)
        except Exception as exc:
            return self._save_alert(
                report,
                recipient,
                ToolStatus.FAILED.value,
                f"高风险预警邮件发送失败：{type(exc).__name__}: {exc}",
            )
        return self._save_alert(report, recipient, ToolStatus.SUCCESS.value, f"高风险预警邮件已发送：reportId={report.id}")

    def _save_alert(self, report: PsychologicalReport, recipient: str, status: str, message: str) -> AlertRecord:
        record = AlertRecord(
            report_id=report.id,
            channel="email",
            recipient=recipient,
            status=status,
            message=message,
        )
        self.db.add(record)
        self.db.commit()
        return record

    def _missing_email_config(self) -> list[str]:
        missing = []
        if not self.settings.smtp_host.strip():
            missing.append("SMTP_HOST")
        if not self._sender():
            missing.append("ALERT_EMAIL_FROM 或 SMTP_USERNAME")
        if not self._recipients():
            missing.append("ALERT_EMAIL_TO")
        return missing

    def _send_alert_email(self, report: PsychologicalReport) -> None:
        message = EmailMessage()
        message["Subject"] = f"{self.settings.alert_email_subject_prefix} reportId={report.id}"
        message["From"] = self._sender()
        message["To"] = ", ".join(self._recipients())
        message.set_content(self._email_body(report))

        context = ssl.create_default_context()
        if self.settings.smtp_use_ssl:
            with smtplib.SMTP_SSL(
                self.settings.smtp_host,
                self.settings.smtp_port,
                timeout=self.settings.smtp_timeout_seconds,
                context=context,
            ) as server:
                self._send_message(server, message)
            return

        with smtplib.SMTP(self.settings.smtp_host, self.settings.smtp_port, timeout=self.settings.smtp_timeout_seconds) as server:
            server.ehlo()
            if self.settings.smtp_use_tls:
                server.starttls(context=context)
                server.ehlo()
            self._send_message(server, message)

    def _send_message(self, server: smtplib.SMTP, message: EmailMessage) -> None:
        if self.settings.smtp_username:
            server.login(self.settings.smtp_username, self.settings.smtp_password)
        server.send_message(message)

    def _email_body(self, report: PsychologicalReport) -> str:
        user = self.db.get(UserAccount, report.user_id)
        username = user.username if user else f"userId={report.user_id}"
        display_name = user.display_name if user else ""
        return "\n".join(
            [
                "SoulFerry 检测到一条高风险心理预警，请尽快安排辅导员或管理员跟进。",
                "",
                f"报告ID：{report.id}",
                f"学生：{display_name} ({username})" if display_name else f"学生：{username}",
                f"风险等级：{report.risk_level}",
                f"情绪标签：{report.emotion}",
                f"置信度：{report.confidence}",
                f"摘要：{report.summary}",
                f"创建时间：{report.created_at.isoformat()}",
                "",
                "学生原始消息：",
                report.content,
            ]
        )

    def _sender(self) -> str:
        return self.settings.alert_email_from.strip() or self.settings.smtp_username.strip()

    def _recipients(self) -> list[str]:
        normalized = self.settings.alert_email_to.replace(";", ",")
        return [recipient.strip() for recipient in normalized.split(",") if recipient.strip()]
